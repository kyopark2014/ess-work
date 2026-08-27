#!/usr/bin/env python3
"""Build an ESS/compliance test-case Excel workbook and publish to S3.

S3 key layout:
    artifacts/{projectName}/{user_id}/tc/{filename}.xlsx

Columns:
    규격명 | 항목 | 원문 | 기준 | 판정결과 | 비고

판정결과 is an Excel dropdown: 합격 / 불합격 / 부분합격

Usage (from application/):
    python skills/testcase-generator/scripts/generate_testcase_xlsx.py \\
        --cases /path/to/cases.json --user ksdyb

JSON schema:
{
  "source_md": "...",
  "title": "...",
  "standard": "NFPA 855 (2023)",
  "cases": [
    {
      "규격명": "NFPA 855 (2023)",
      "항목": "4.2.1.1",
      "원문": "...",
      "기준": "ESS 설치·교체·커미셔닝·사용 관련 도면·시방서가 AHJ에 제출·승인되었고, (1)~(8) 항목이 누락 없이 포함되어 있는지 확인",
      "판정결과": "",
      "비고": ""
    }
  ]
}
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib import parse


COLUMNS = ("규격명", "항목", "원문", "기준", "판정결과", "비고")
JUDGMENT_OPTIONS = ("합격", "불합격", "부분합격")
COLUMN_ALIASES = {
    "규격명": ("규격명", "standard", "standard_name", "규격"),
    "항목": ("항목", "item", "clause", "id"),
    "원문": ("원문", "source", "text", "requirement"),
    "기준": ("기준", "criteria", "pass_fail_criteria"),
    "판정결과": ("판정결과", "result", "judgment", "verdict"),
    "비고": ("비고", "notes", "remark", "remarks", "comment"),
}


def _application_dir() -> Path:
    # .../application/skills/testcase-generator/scripts/this.py
    return Path(__file__).resolve().parents[3]


def _load_config() -> dict[str, Any]:
    cfg_path = _application_dir() / "config.json"
    if not cfg_path.is_file():
        return {}
    try:
        return json.loads(cfg_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _session_storage() -> Path:
    env = os.environ.get("SESSION_STORAGE_DIR")
    if env:
        return Path(env)
    return _application_dir() / ".session_storage"


def _safe_user(user_id: str | None) -> str:
    raw = (user_id or "").strip() or "default"
    return (
        raw.replace("/", "_").replace("\\", "_").replace("..", "_")[:128] or "default"
    )


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook  # noqa: F401
    except ImportError:
        import subprocess

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl"],
            stdout=subprocess.DEVNULL,
        )


def _resolve_standard(data: dict[str, Any]) -> str:
    for key in ("standard", "규격명", "규격", "standard_name"):
        value = data.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _infer_standard_from_source(source_md: str) -> str:
    """Best-effort standard label from filename stem when JSON omits it."""
    stem = Path(source_md or "").stem.strip()
    if not stem:
        return ""
    upper = stem.upper().replace("-", "_")
    year_m = re.search(r"(20\d{2})", stem)
    year = year_m.group(1) if year_m else ""
    if upper.startswith("NFPA"):
        # NFPA855_2023 / NFPA_855_2020 → NFPA 855 (2023)
        m = re.match(r"NFPA[_]?(\d+)", upper)
        if m:
            label = f"NFPA {m.group(1)}"
            return f"{label} ({year})" if year else label
    if "9540" in upper or upper.startswith("S9540") or "UL9540" in upper.replace("_", ""):
        label = "UL 9540"
        return f"{label} ({year})" if year else label
    return stem.replace("_", " ")


def _strip_standard_prefix(item: str, standard: str) -> str:
    """If 항목 still contains '규격명 …', peel it so 항목 is clause-only."""
    clause = (item or "").strip()
    name = (standard or "").strip()
    if not clause or not name:
        return clause
    norm_clause = re.sub(r"\s+", " ", clause)
    norm_name = re.sub(r"\s+", " ", name)
    if norm_clause.casefold() == norm_name.casefold():
        return ""
    if norm_clause.casefold().startswith(norm_name.casefold() + " "):
        return norm_clause[len(norm_name) :].strip()
    # Also peel common variants without parentheses year: "NFPA 855 1.3"
    loose = re.sub(r"\s*\(\d{4}\)\s*", " ", norm_name).strip()
    if loose and norm_clause.casefold().startswith(loose.casefold() + " "):
        return norm_clause[len(loose) :].strip()
    return clause


def _cell_value(row: dict[str, Any], col: str) -> str:
    for key in COLUMN_ALIASES[col]:
        if key in row and row[key] is not None:
            return str(row[key]).strip()
    return ""


def _normalize_judgment(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    if text in JUDGMENT_OPTIONS:
        return text
    # Tolerate common variants
    mapping = {
        "pass": "합격",
        "fail": "불합격",
        "partial": "부분합격",
        "partial pass": "부분합격",
        "부분 합격": "부분합격",
    }
    return mapping.get(text.casefold(), text if text in JUDGMENT_OPTIONS else "")


def _load_cases(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("cases JSON must be an object")
    cases = data.get("cases")
    if not isinstance(cases, list) or not cases:
        raise ValueError("cases JSON must include a non-empty 'cases' array")

    standard = _resolve_standard(data)
    if not standard:
        standard = _infer_standard_from_source(str(data.get("source_md") or ""))
    if standard:
        data["standard"] = standard

    normalized: list[dict[str, str]] = []
    for i, row in enumerate(cases):
        if not isinstance(row, dict):
            raise ValueError(f"cases[{i}] must be an object")
        item = {col: _cell_value(row, col) for col in COLUMNS}

        # Per-row 규격명 overrides root standard when present.
        row_standard = item["규격명"] or standard
        item["규격명"] = row_standard
        item["항목"] = _strip_standard_prefix(item["항목"], row_standard)
        item["판정결과"] = _normalize_judgment(item["판정결과"])

        if not item["항목"] and not item["원문"]:
            raise ValueError(f"cases[{i}] needs at least 항목 or 원문")
        if not item["규격명"]:
            raise ValueError(
                f"cases[{i}] needs 규격명 (set root 'standard' or per-row 규격명)"
            )
        normalized.append(item)
    data["cases"] = normalized
    return data


def _default_output_name(payload: dict[str, Any]) -> str:
    source = str(payload.get("source_md") or "").strip()
    if source:
        stem = Path(source).stem
        if stem:
            return f"{stem}_testcase.xlsx"
    title = str(payload.get("title") or "testcase").strip()
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in title)[:80]
    return f"{safe or 'testcase'}.xlsx"


def _write_xlsx(payload: dict[str, Any], dest: Path) -> int:
    _ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    cases = payload["cases"]
    last_row = max(len(cases) + 1, 2)
    for row_idx, row in enumerate(cases, start=2):
        for col_idx, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(name, ""))
            cell.font = body_font
            cell.border = thin
            if name in ("규격명", "항목", "판정결과"):
                cell.alignment = center
            else:
                cell.alignment = wrap

    # 규격명 | 항목 | 원문 | 기준 | 판정결과 | 비고
    widths = (22, 14, 64, 36, 14, 28)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = 60

    # Dropdown for 판정결과 (column E)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(JUDGMENT_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,  # False => show the arrow in Excel
        showErrorMessage=True,
        errorTitle="판정결과",
        error="합격, 불합격, 부분합격 중에서 선택하세요.",
        promptTitle="판정결과",
        prompt="합격 / 불합격 / 부분합격",
    )
    # Reserve extra empty rows so users can fill more later.
    dv_end = max(last_row, 500)
    dv.add(f"E2:E{dv_end}")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{last_row}"

    # Meta sheet
    meta = wb.create_sheet("Meta", 1)
    meta["A1"] = "title"
    meta["B1"] = payload.get("title") or ""
    meta["A2"] = "standard"
    meta["B2"] = payload.get("standard") or ""
    meta["A3"] = "source_md"
    meta["B3"] = payload.get("source_md") or ""
    meta["A4"] = "generated_at"
    meta["B4"] = datetime.now(timezone.utc).isoformat()
    meta["A5"] = "row_count"
    meta["B5"] = len(cases)
    meta["A6"] = "judgment_options"
    meta["B6"] = ", ".join(JUDGMENT_OPTIONS)
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 80
    for r in range(1, 7):
        meta.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10)
        meta.cell(row=r, column=2).font = Font(name="Arial", size=10)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return len(cases)


def _upload_to_s3(
    local_path: Path,
    *,
    user_id: str,
    file_name: str,
    config: dict[str, Any],
) -> dict[str, Any]:
    bucket = (config.get("s3_bucket") or "").strip()
    region = (config.get("region") or "us-west-2").strip()
    project = (config.get("projectName") or "default").strip().strip("/") or "default"
    sharing = (config.get("sharing_url") or "").rstrip("/")
    segment = _safe_user(user_id)
    s3_key = f"artifacts/{project}/{segment}/tc/{file_name}"
    result = {
        "s3_key": s3_key,
        "url": None,
        "uploaded": False,
    }
    if sharing:
        parts = [parse.quote(p) for p in s3_key.split("/")]
        result["url"] = f"{sharing}/{'/'.join(parts)}"

    if not bucket:
        result["error"] = "s3_bucket is not configured in config.json"
        return result

    try:
        import boto3
    except ImportError:
        result["error"] = "boto3 is not installed"
        return result

    try:
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        client = boto3.client("s3", region_name=region)
        with local_path.open("rb") as f:
            client.put_object(
                Bucket=bucket,
                Key=s3_key,
                Body=f.read(),
                ContentType=content_type,
                ContentDisposition=f'attachment; filename="{file_name}"',
                CacheControl="no-cache, max-age=0, must-revalidate",
            )
        result["uploaded"] = True
        return result
    except Exception as exc:
        result["error"] = str(exc)
        return result


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate test-case xlsx and upload to artifacts/{project}/{user}/tc/"
    )
    parser.add_argument("--cases", required=True, help="Path to cases JSON")
    parser.add_argument("--user", default=None, help="User id for S3 path segment")
    parser.add_argument(
        "--output-name",
        default=None,
        help="Output xlsx file name (default: {md_stem}_testcase.xlsx)",
    )
    parser.add_argument(
        "--local-only",
        action="store_true",
        help="Skip S3 upload; write local file only",
    )
    args = parser.parse_args()

    cases_path = Path(os.path.expanduser(args.cases)).resolve()
    if not cases_path.is_file():
        print(json.dumps({"ok": False, "error": f"cases file not found: {cases_path}"}))
        return 1

    try:
        payload = _load_cases(cases_path)
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1

    config = _load_config()
    user_id = _safe_user(
        args.user
        or os.environ.get("ESS_USER_ID")
        or os.environ.get("USER_ID")
        or config.get("default_user")
    )
    file_name = (args.output_name or _default_output_name(payload)).strip()
    if not file_name.lower().endswith(".xlsx"):
        file_name = f"{file_name}.xlsx"
    file_name = os.path.basename(file_name)

    local_dir = _session_storage() / user_id / "artifacts" / "tc"
    local_path = local_dir / file_name

    try:
        rows = _write_xlsx(payload, local_path)
    except Exception as exc:
        print(json.dumps({"ok": False, "error": f"xlsx write failed: {exc}"}))
        return 1

    upload: dict[str, Any] = {
        "s3_key": None,
        "url": None,
        "uploaded": False,
    }
    if not args.local_only:
        upload = _upload_to_s3(
            local_path, user_id=user_id, file_name=file_name, config=config
        )

    out = {
        "ok": True,
        "local_path": str(local_path),
        "s3_key": upload.get("s3_key"),
        "url": upload.get("url"),
        "uploaded": bool(upload.get("uploaded")),
        "rows": rows,
        "user_id": user_id,
        "source_md": payload.get("source_md"),
        "title": payload.get("title"),
        "standard": payload.get("standard"),
        "columns": list(COLUMNS),
        "judgment_options": list(JUDGMENT_OPTIONS),
    }
    if upload.get("error"):
        out["upload_error"] = upload["error"]
    print(json.dumps(out, ensure_ascii=False, indent=2))
    return 0 if out["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
