#!/usr/bin/env python3
"""Build a regulation compliance Excel report from an evaluation JSON.

Sheets (fixed order):
  1. Summary
  2. 판정결과_상세
  3. 조치필요_항목
  4. coverage sheet (name from meta.coverage_sheet_name)

Usage (from application/):
    python skills/regulation-evaluator/scripts/generate_compliance_report.py \\
        --evaluation /path/to/evaluation.json \\
        --output-name NFPA855_F2XX_ERG_Compliance_Report.xlsx

Evaluation JSON schema: see ../references/report-format.md
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any


VERDICTS = ("합격", "부분합격", "불합격", "확인불가", "해당없음")
VERDICT_DESC = {
    "합격": "대상 문서에서 요구사항 충족 확인",
    "부분합격": "일부 요구사항은 확인되나 나머지는 추가 문서 필요",
    "불합격": "요구사항 미충족 확인",
    "확인불가": "문서 범위 초과 – 현장·설계문서 별도 확인 필요",
    "해당없음": "해당 조항 적용 제외 대상",
}
PRIORITY_LABEL = {
    "높음": "⭐⭐⭐ 높음",
    "중간": "⭐⭐ 중간",
    "낮음": "⭐ 낮음",
}
COVERAGE_NORMALIZE = {
    "잘 커버": "🟢 잘 커버",
    "부분 커버": "🔵 부분 커버",
    "미커버": "🔴 미커버",
    "🟢 잘 커버": "🟢 잘 커버",
    "🔵 부분 커버": "🔵 부분 커버",
    "🔴 미커버": "🔴 미커버",
}

COLOR = {
    "header_bg": "1F4E79",
    "title_bg": "1F4E79",
    "cov_title_bg": "375623",
    "pass_bg": "C6EFCE",
    "pass_font": "006100",
    "partial_bg": "FFEB9C",
    "partial_font": "9C5700",
    "fail_bg": "FFC7CE",
    "fail_font": "9C0006",
    "unknown_bg": "D9D9D9",
    "unknown_font": "595959",
    "na_bg": "DDEBF7",
    "na_font": "1F4E79",
    "chapter_bg": "D6DCE4",
    "alt_row": "F5F5F5",
    "white": "FFFFFF",
    "black": "000000",
    "action_title_bg": "C65911",
}

FONT_NAME = "맑은 고딕"


def _application_dir() -> Path:
    # .../application/skills/regulation-evaluator/scripts/this.py
    return Path(__file__).resolve().parents[3]


def _session_storage() -> Path:
    env = os.environ.get("SESSION_STORAGE_DIR")
    if env:
        return Path(env)
    return _application_dir() / ".session_storage"


def _artifacts_dir(user_id: str | None) -> Path:
    env = os.environ.get("ARTIFACTS_DIR")
    if env:
        return Path(env)
    uid = (user_id or os.environ.get("USER_ID") or "default").strip() or "default"
    uid = uid.replace("/", "_").replace("\\", "_").replace("..", "_")[:128]
    return _session_storage() / uid / "artifacts"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import subprocess

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl"],
            stdout=subprocess.DEVNULL,
        )


def _first(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _pick(obj: dict[str, Any], *keys: str) -> str:
    for key in keys:
        if key in obj and obj[key] is not None:
            text = str(obj[key]).strip()
            if text:
                return text
    return ""


def _as_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]
    text = str(value).strip()
    return [text] if text else []


def _infer_chapter(item: str) -> str:
    text = (item or "").strip()
    m = re.match(r"^(\d+)", text)
    if not m:
        return ""
    return f"Chapter {m.group(1)}"


def _guess_coverage_label(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return "🔵 부분 커버"
    mapped = COVERAGE_NORMALIZE.get(text)
    if mapped:
        return mapped
    lower = text.lower()
    if any(k in text for k in ("미커버", "미포함", "범위 밖", "없음", "누락")) or "missing" in lower:
        return "🔴 미커버"
    if any(k in text for k in ("잘 커버", "합격", "충분", "핵심")) or "pass" in lower:
        return "🟢 잘 커버"
    if any(k in text for k in ("부분", "미첨부", "확인", "권장", "근거")):
        return "🔵 부분 커버"
    return text if text.startswith(("🟢", "🔵", "🔴")) else f"🔵 {text}"


def _default_priority_for_verdict(verdict: str) -> str:
    if verdict == "불합격":
        return "높음"
    if verdict == "부분합격":
        return "중간"
    if verdict == "확인불가":
        return "낮음"
    return "중간"


def _normalize_meta(raw: dict[str, Any] | None) -> dict[str, Any]:
    meta = dict(raw or {})
    standard = _pick(meta, "standard", "규격명", "검토 기준")
    target = _pick(
        meta,
        "target_doc",
        "target_document",
        "document",
        "doc_name",
        "검토 대상 문서",
        "대상문서",
    )
    product = _pick(meta, "product", "ess_product", "ESS 제품", "제품")
    battery = _pick(meta, "battery_type", "battery", "배터리 타입", "배터리")
    review_date = _pick(meta, "review_date", "evaluated_at", "date", "검토 일자")
    if review_date in {"auto", "today", "now"}:
        review_date = date.today().isoformat()
    scope = _pick(
        meta,
        "scope",
        "excluded_note",
        "doc_type",
        "검토 범위",
        "범위",
    )
    title = _pick(meta, "title", "report_title")
    if not title:
        std_short = standard or "규격"
        doc_short = target or "대상 문서"
        title = f"{std_short} 적합성 검토 리포트 — {doc_short}"

    doc_type = _pick(meta, "doc_type", "document_type")
    coverage_sheet = _pick(meta, "coverage_sheet_name") or "문서_커버리지_매핑"
    coverage_title = _pick(meta, "coverage_title") or "문서 섹션 ↔ 규격 조항 커버리지 매핑"
    evidence_label = _pick(meta, "evidence_col_label") or "판정 근거 (대상 문서 근거)"
    action_req = _pick(meta, "action_req_label") or "규격 요구사항"
    action_status = _pick(meta, "action_status_label") or "현재 문서 상태"

    return {
        **meta,
        "title": title,
        "target_doc": target,
        "standard": standard,
        "product": product,
        "battery_type": battery,
        "review_date": review_date or date.today().isoformat(),
        "scope": scope,
        "doc_type": doc_type,
        "output_name": _pick(meta, "output_name"),
        "coverage_sheet_name": coverage_sheet[:31],
        "coverage_title": coverage_title,
        "coverage_section_header": _pick(meta, "coverage_section_header") or "문서 섹션",
        "coverage_clause_header": _pick(meta, "coverage_clause_header") or "관련 규격 조항",
        "evidence_col_label": evidence_label,
        "action_req_label": action_req,
        "action_status_label": action_status,
    }


def _normalize_summary(
    raw: dict[str, Any] | None, results: list[dict[str, Any]]
) -> dict[str, Any]:
    summary = dict(raw or {})
    strengths = _as_list(
        summary.get("strengths")
        or summary.get("강점")
        or summary.get("highlights")
    )
    gaps = _as_list(
        summary.get("gaps")
        or summary.get("improvements")
        or summary.get("보완")
        or summary.get("보완필요")
        or summary.get("actions_summary")
    )
    needs = _as_list(
        summary.get("needs_verification")
        or summary.get("needs_check")
        or summary.get("별도확인")
        or summary.get("out_of_scope")
    )
    scope_note = _pick(summary, "scope_note", "note", "범위설명")

    if not strengths:
        strengths = [
            f"{r.get('항목')}: {(r.get('판정근거') or '')[:80]}"
            for r in results
            if r.get("판정결과") == "합격"
        ][:5]
    if not gaps:
        gaps = [
            f"{r.get('항목')}: {(r.get('비고') or r.get('판정근거') or '')[:100]}"
            for r in results
            if r.get("판정결과") in {"부분합격", "불합격"}
        ][:8]
    if not needs:
        needs = [
            f"{r.get('항목')}: {(r.get('비고') or r.get('판정근거') or '')[:100]}"
            for r in results
            if r.get("판정결과") == "확인불가"
        ][:8]

    return {
        **summary,
        "strengths": strengths,
        "gaps": gaps,
        "needs_verification": needs,
        "scope_note": scope_note,
    }


def _normalize_results(
    raw_results: list[Any], standard: str
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in raw_results:
        if not isinstance(row, dict):
            continue
        item = _pick(row, "항목", "item", "clause", "id")
        evidence = _pick(
            row,
            "판정근거",
            "판정 근거",
            "근거",
            "evidence",
            "rationale",
            "reason",
        )
        note = _pick(row, "비고", "notes", "remark", "remarks", "comment")
        # If evidence empty but note looks like evidence, keep note as note only.
        chapter = _pick(row, "chapter", "장", "Chapter") or _infer_chapter(item)
        out.append(
            {
                "chapter": chapter,
                "규격명": _pick(row, "규격명", "standard") or standard,
                "항목": item,
                "원문": _pick(row, "원문", "원문 (요약)", "source", "text", "requirement"),
                "기준": _pick(row, "기준", "판정 기준", "criteria"),
                "판정결과": _pick(row, "판정결과", "result", "verdict", "judgment"),
                "판정근거": evidence,
                "비고": note,
                "요약": _pick(row, "요약", "summary", "title"),
            }
        )
    return out


def _normalize_actions(
    raw_actions: list[Any], results: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    by_item = {str(r.get("항목") or "").strip(): r for r in results}
    out: list[dict[str, Any]] = []

    for row in raw_actions or []:
        if not isinstance(row, dict):
            continue
        item = _pick(row, "항목", "item", "clause")
        matched = by_item.get(item, {})
        requirement = _pick(
            row,
            "요구사항",
            "규격 요구사항",
            "requirement",
            "기준",
            "요약",
        ) or _pick(matched, "기준", "요약", "원문")
        status = _pick(
            row,
            "현재상태",
            "현재 문서 상태",
            "status",
            "근거",
            "판정근거",
            "evidence",
        ) or _pick(matched, "판정근거")
        action = _pick(
            row,
            "권고조치",
            "권고 조치사항",
            "조치사항",
            "action",
            "recommendation",
            "비고",
        ) or _pick(matched, "비고")
        priority = _pick(row, "우선순위", "priority") or _default_priority_for_verdict(
            _pick(row, "판정결과") or _pick(matched, "판정결과")
        )
        if not any([item, requirement, status, action]):
            continue
        out.append(
            {
                "항목": item,
                "요구사항": requirement,
                "현재상태": status,
                "권고조치": action,
                "우선순위": priority,
            }
        )

    if out:
        return out

    # Auto-build actions from non-pass results when actions omitted.
    for r in results:
        verdict = str(r.get("판정결과") or "").strip()
        if verdict not in {"부분합격", "불합격", "확인불가"}:
            continue
        out.append(
            {
                "항목": r.get("항목") or "",
                "요구사항": r.get("기준") or r.get("요약") or (r.get("원문") or "")[:120],
                "현재상태": r.get("판정근거") or "",
                "권고조치": r.get("비고") or "추가 문서·현장 확인 필요",
                "우선순위": _default_priority_for_verdict(verdict),
            }
        )
    return out


def _normalize_coverage_rows(raw_coverage: list[Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in raw_coverage or []:
        if not isinstance(row, dict):
            continue
        section = _pick(
            row,
            "섹션",
            "문서_섹션",
            "문서섹션",
            "section",
            "doc_section",
            "ERG 섹션",
        )
        content = _pick(
            row,
            "주요내용",
            "주요 내용",
            "내용",
            "content",
            "summary",
            "평가설명",
        )
        clauses = _pick(
            row,
            "관련조항",
            "관련 조항",
            "대응_조항",
            "대응조항",
            "clauses",
            "nfpa",
        )
        coverage_raw = _pick(
            row,
            "커버리지",
            "커버리지 평가",
            "평가",
            "coverage",
            "rating",
        )
        # If "주요내용" missing, reuse free-text evaluation as content.
        if not content and coverage_raw and coverage_raw not in COVERAGE_NORMALIZE:
            content = coverage_raw
        if not any([section, content, clauses]):
            continue
        out.append(
            {
                "섹션": section,
                "주요내용": content,
                "관련조항": clauses,
                "커버리지": _guess_coverage_label(coverage_raw),
            }
        )
    return out


def _normalize_evaluation(data: dict[str, Any]) -> dict[str, Any]:
    meta = _normalize_meta(data.get("meta") if isinstance(data.get("meta"), dict) else {})
    results = _normalize_results(
        data.get("results") if isinstance(data.get("results"), list) else [],
        meta.get("standard") or "",
    )
    if not results:
        raise ValueError("evaluation JSON 'results' must be a non-empty list")
    summary = _normalize_summary(
        data.get("summary") if isinstance(data.get("summary"), dict) else {},
        results,
    )
    if summary.get("scope_note") and not meta.get("scope"):
        meta["scope"] = summary["scope_note"]
    actions = _normalize_actions(
        data.get("actions") if isinstance(data.get("actions"), list) else [],
        results,
    )
    coverage = _normalize_coverage_rows(
        data.get("coverage") if isinstance(data.get("coverage"), list) else []
    )
    return {
        "meta": meta,
        "summary": summary,
        "results": results,
        "actions": actions,
        "coverage": coverage,
    }


def _load_evaluation(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("evaluation JSON must be an object")
    if "meta" not in data or "results" not in data:
        raise ValueError("evaluation JSON missing 'meta' or 'results'")
    return _normalize_evaluation(data)


def _font(color: str = "000000", *, bold: bool = False, size: int = 10):
    from openpyxl.styles import Font

    return Font(name=FONT_NAME, color=color, bold=bold, size=size)


def _fill(hex_color: str):
    from openpyxl.styles import PatternFill

    return PatternFill("solid", fgColor=hex_color)


def _align(horizontal: str = "left", *, wrap: bool = True, vertical: str = "center"):
    from openpyxl.styles import Alignment

    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _border():
    from openpyxl.styles import Border, Side

    side = Side(style="thin", color="B0B0B0")
    return Border(left=side, right=side, top=side, bottom=side)


def _verdict_style(verdict: str) -> tuple[str, str]:
    v = (verdict or "").strip()
    if v == "합격":
        return COLOR["pass_bg"], COLOR["pass_font"]
    if v == "부분합격":
        return COLOR["partial_bg"], COLOR["partial_font"]
    if v == "불합격":
        return COLOR["fail_bg"], COLOR["fail_font"]
    if v == "확인불가":
        return COLOR["unknown_bg"], COLOR["unknown_font"]
    if v == "해당없음":
        return COLOR["na_bg"], COLOR["na_font"]
    return COLOR["white"], COLOR["black"]


def _coverage_style(label: str) -> tuple[str, str]:
    if "잘 커버" in label:
        return COLOR["pass_bg"], COLOR["pass_font"]
    if "부분" in label:
        return COLOR["partial_bg"], COLOR["partial_font"]
    if "미커버" in label:
        return COLOR["fail_bg"], COLOR["fail_font"]
    return COLOR["white"], COLOR["black"]


def _normalize_priority(raw: str) -> str:
    text = (raw or "").strip()
    for key, label in PRIORITY_LABEL.items():
        if key in text or label == text:
            return label
    if text.startswith("⭐"):
        return text
    return PRIORITY_LABEL.get(text, text or PRIORITY_LABEL["중간"])


def _normalize_coverage(raw: str) -> str:
    text = (raw or "").strip()
    return COVERAGE_NORMALIZE.get(text, text)


def _count_verdicts(results: list[dict[str, Any]]) -> Counter:
    c: Counter = Counter()
    for row in results:
        v = str(row.get("판정결과") or "").strip()
        if v in VERDICTS:
            c[v] += 1
    return c


def _bullet_join(items: list[Any] | str | None) -> str:
    if items is None:
        return ""
    if isinstance(items, str):
        return items
    lines = [str(x).strip() for x in items if str(x).strip()]
    return "\n".join(f"• {x}" for x in lines)


def _default_output_name(meta: dict[str, Any], eval_path: Path) -> str:
    name = str(meta.get("output_name") or "").strip()
    if name:
        return name if name.endswith(".xlsx") else f"{name}.xlsx"
    stem = eval_path.stem.replace("_evaluation", "")
    return f"{stem or 'Compliance'}_Compliance_Report.xlsx"


def _unique_path(path: Path) -> Path:
    """If path exists, return path with _1, _2, … before the suffix."""
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    n = 1
    while True:
        candidate = path.with_name(f"{stem}_{n}{suffix}")
        if not candidate.exists():
            return candidate
        n += 1


def _write_summary(wb, data: dict[str, Any]):
    from openpyxl.utils import get_column_letter

    meta = data["meta"]
    summary = data.get("summary") or {}
    counts = _count_verdicts(data["results"])
    total = sum(counts.values()) or 1

    ws = wb.create_sheet("Summary", 0)
    ws.merge_cells("A1:G1")
    ws["A1"] = meta.get("title") or "적합성 검토 리포트"
    ws["A1"].fill = _fill(COLOR["title_bg"])
    ws["A1"].font = _font("FFFFFF", bold=True, size=13)
    ws["A1"].alignment = _align("center", wrap=True)
    ws.row_dimensions[1].height = 28

    meta_rows = [
        ("검토 대상 문서", meta.get("target_doc") or ""),
        ("검토 기준", meta.get("standard") or ""),
        ("ESS 제품", meta.get("product") or ""),
        ("배터리 타입", meta.get("battery_type") or ""),
        ("검토 일자", meta.get("review_date") or date.today().isoformat()),
        ("검토 범위", meta.get("scope") or ""),
    ]
    for i, (label, value) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True, size=10)
        ws.cell(row=i, column=1).fill = _fill("E7E6E6")
        ws.cell(row=i, column=2, value=value).font = _font(size=10)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = _border()
            ws.cell(row=i, column=col).alignment = _align("left", wrap=True)

    ws.cell(row=9, column=1, value="📊 판정 결과 요약").font = _font(bold=True, size=12)
    headers = ["판정결과", "건수", "비율", "설명"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=h)
        cell.fill = _fill(COLOR["header_bg"])
        cell.font = _font("FFFFFF", bold=True, size=10)
        cell.alignment = _align("center", wrap=False)
        cell.border = _border()

    for i, verdict in enumerate(VERDICTS, start=11):
        n = counts.get(verdict, 0)
        pct = f"{(n / total) * 100:.1f}%"
        bg, fg = _verdict_style(verdict)
        values = [verdict, n, pct, VERDICT_DESC[verdict]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _border()
            cell.alignment = _align("center" if col < 4 else "left", wrap=True)
            if col == 1:
                cell.fill = _fill(bg)
                cell.font = _font(fg, bold=True, size=10)
            else:
                cell.font = _font(size=10)

    sum_row = 16
    ws.cell(row=sum_row, column=1, value="합계").font = _font(bold=True, size=10)
    ws.cell(row=sum_row, column=2, value=sum(counts.values())).font = _font(
        bold=True, size=10
    )
    for col in range(1, 5):
        ws.cell(row=sum_row, column=col).border = _border()
        ws.cell(row=sum_row, column=col).fill = _fill("E7E6E6")

    ws.cell(row=18, column=1, value="🔍 주요 발견사항").font = _font(bold=True, size=12)
    findings = [
        ("✅ 강점", _bullet_join(summary.get("strengths"))),
        ("⚠️ 보완 필요", _bullet_join(summary.get("gaps"))),
        ("❓ 별도 확인 필요", _bullet_join(summary.get("needs_verification"))),
    ]
    for i, (label, text) in enumerate(findings, start=19):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True, size=10)
        ws.cell(row=i, column=1).fill = _fill("E7E6E6")
        ws.cell(row=i, column=1).alignment = _align("left", wrap=True)
        ws.cell(row=i, column=2, value=text).font = _font(size=9)
        ws.cell(row=i, column=2).alignment = _align("left", wrap=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        ws.row_dimensions[i].height = max(48, 14 * (text.count("\n") + 2))
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = _border()

    widths = [22, 55, 12, 55, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_detail(wb, data: dict[str, Any]):
    from openpyxl.utils import get_column_letter

    meta = data["meta"]
    evidence_label = (
        meta.get("evidence_col_label") or "판정 근거 (대상 문서 근거)"
    )
    headers = [
        "규격명",
        "항목",
        "원문 (요약)",
        "판정 기준",
        "판정결과",
        evidence_label,
        "비고",
    ]
    widths = [18, 14, 42, 32, 12, 42, 28]

    ws = wb.create_sheet("판정결과_상세")
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _fill(COLOR["header_bg"])
        cell.font = _font("FFFFFF", bold=True, size=10)
        cell.alignment = _align("center", wrap=False)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    row_idx = 2
    last_chapter = None
    for item in data["results"]:
        chapter = str(item.get("chapter") or "").strip()
        if chapter and chapter != last_chapter:
            label = chapter if chapter.lower().startswith("chapter") else f"Chapter {chapter}"
            ws.merge_cells(
                start_row=row_idx, start_column=1, end_row=row_idx, end_column=7
            )
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.fill = _fill(COLOR["chapter_bg"])
            cell.font = _font(bold=True, size=10)
            cell.alignment = _align("left", wrap=False)
            cell.border = _border()
            for col in range(2, 8):
                c = ws.cell(row=row_idx, column=col)
                c.fill = _fill(COLOR["chapter_bg"])
                c.border = _border()
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1
            last_chapter = chapter

        verdict = str(item.get("판정결과") or "").strip()
        bg, fg = _verdict_style(verdict)
        source = item.get("원문") or item.get("원문 (요약)") or ""
        source = str(source).replace("\\*", "*").replace("\\n", "\n")
        note = item.get("비고") or ""
        if not str(note).strip():
            note = "—"
        values = [
            item.get("규격명") or meta.get("standard") or "",
            item.get("항목") or "",
            source,
            item.get("기준") or item.get("판정 기준") or "",
            verdict,
            item.get("판정근거") or item.get("판정 근거") or "",
            note,
        ]
        alt = COLOR["alt_row"] if row_idx % 2 == 0 else COLOR["white"]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _border()
            if col == 5:
                cell.fill = _fill(bg)
                cell.font = _font(fg, bold=True, size=9)
                cell.alignment = _align("center", wrap=False)
            else:
                cell.fill = _fill(alt)
                cell.font = _font(size=9)
                cell.alignment = _align(
                    "center" if col in (1, 2) else "left", wrap=True
                )
        ws.row_dimensions[row_idx].height = 55
        row_idx += 1


def _write_actions(wb, data: dict[str, Any]):
    from openpyxl.utils import get_column_letter

    meta = data["meta"]
    req_label = meta.get("action_req_label") or "규격 요구사항"
    status_label = meta.get("action_status_label") or "현재 문서 상태"
    headers = ["항목", req_label, status_label, "권고 조치사항", "우선순위"]
    widths = [14, 36, 36, 40, 14]

    ws = wb.create_sheet("조치필요_항목")
    ws.merge_cells("A1:E1")
    ws["A1"] = "🔧 보완 및 조치 필요 항목 (Action Required)"
    ws["A1"].fill = _fill(COLOR["action_title_bg"])
    ws["A1"].font = _font("FFFFFF", bold=True, size=12)
    ws["A1"].alignment = _align("center", wrap=False)
    ws.row_dimensions[1].height = 24

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = _fill(COLOR["header_bg"])
        cell.font = _font("FFFFFF", bold=True, size=10)
        cell.alignment = _align("center", wrap=False)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"

    actions = data.get("actions") or []
    if not actions:
        ws.cell(row=3, column=1, value="(조치 필요 항목 없음)")
        ws.cell(row=3, column=1).font = _font(size=10)
        return

    for i, item in enumerate(actions, start=3):
        priority = _normalize_priority(str(item.get("우선순위") or ""))
        values = [
            item.get("항목") or "",
            item.get("요구사항") or "",
            item.get("현재상태") or "",
            item.get("권고조치") or item.get("권고 조치사항") or "",
            priority,
        ]
        alt = COLOR["alt_row"] if i % 2 == 0 else COLOR["white"]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _border()
            cell.font = _font(size=9)
            if col == 5:
                if "높음" in priority:
                    cell.fill = _fill(COLOR["fail_bg"])
                    cell.font = _font(COLOR["fail_font"], bold=True, size=9)
                elif "중간" in priority:
                    cell.fill = _fill(COLOR["partial_bg"])
                    cell.font = _font(COLOR["partial_font"], bold=True, size=9)
                else:
                    cell.fill = _fill(COLOR["na_bg"])
                    cell.font = _font(COLOR["na_font"], bold=True, size=9)
                cell.alignment = _align("center", wrap=False)
            else:
                cell.fill = _fill(alt)
                cell.alignment = _align("left", wrap=True)
        ws.row_dimensions[i].height = 50


def _write_coverage(wb, data: dict[str, Any]):
    from openpyxl.utils import get_column_letter

    meta = data["meta"]
    sheet_name = str(meta.get("coverage_sheet_name") or "문서_커버리지_매핑")[:31]
    title = meta.get("coverage_title") or "문서 섹션 ↔ 규격 조항 커버리지 매핑"
    section_header = meta.get("coverage_section_header") or "문서 섹션"
    clause_header = meta.get("coverage_clause_header") or "관련 규격 조항"

    headers = [section_header, "주요 내용", clause_header, "커버리지 평가"]
    widths = [20, 55, 35, 16]

    ws = wb.create_sheet(sheet_name)
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].fill = _fill(COLOR["cov_title_bg"])
    ws["A1"].font = _font("FFFFFF", bold=True, size=13)
    ws["A1"].alignment = _align("center", wrap=False)
    ws.row_dimensions[1].height = 26

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = _fill(COLOR["header_bg"])
        cell.font = _font("FFFFFF", bold=True, size=10)
        cell.alignment = _align("center", wrap=False)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"

    coverage = data.get("coverage") or []
    for i, item in enumerate(coverage, start=3):
        cov = _guess_coverage_label(str(item.get("커버리지") or ""))
        bg, fg = _coverage_style(cov)
        values = [
            item.get("섹션") or "",
            item.get("주요내용") or item.get("주요 내용") or "",
            item.get("관련조항") or item.get("관련 조항") or "",
            cov,
        ]
        alt = COLOR["alt_row"] if i % 2 == 0 else COLOR["white"]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _border()
            if col == 4:
                cell.fill = _fill(bg)
                cell.font = _font(fg, bold=True, size=9)
                cell.alignment = _align("center", wrap=False)
            else:
                cell.fill = _fill(alt)
                cell.font = _font(size=9)
                cell.alignment = _align("left" if col > 1 else "center", wrap=True)
        ws.row_dimensions[i].height = 60


def build_workbook(data: dict[str, Any]):
    _ensure_openpyxl()
    from openpyxl import Workbook

    wb = Workbook()
    # remove default sheet; _write_summary inserts Summary at index 0
    default = wb.active
    wb.remove(default)

    _write_summary(wb, data)
    _write_detail(wb, data)
    _write_actions(wb, data)
    _write_coverage(wb, data)

    # Ensure Summary is first
    if "Summary" in wb.sheetnames and wb.sheetnames[0] != "Summary":
        wb.move_sheet("Summary", offset=-wb.sheetnames.index("Summary"))
    return wb


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate compliance Excel report from evaluation JSON"
    )
    parser.add_argument("--evaluation", required=True, help="Path to evaluation JSON")
    parser.add_argument(
        "--output-name",
        default=None,
        help="Output xlsx file name under ARTIFACTS_DIR/reports/",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Absolute output path (overrides --output-name and ARTIFACTS_DIR)",
    )
    parser.add_argument(
        "--user",
        default=None,
        help="User id for default ARTIFACTS_DIR resolution",
    )
    args = parser.parse_args()

    eval_path = Path(os.path.expanduser(args.evaluation)).resolve()
    if not eval_path.is_file():
        print(json.dumps({"ok": False, "error": f"evaluation not found: {eval_path}"}))
        return 1

    try:
        data = _load_evaluation(eval_path)
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1

    meta = data["meta"]
    if args.output:
        out_path = Path(os.path.expanduser(args.output)).resolve()
    else:
        name = args.output_name or _default_output_name(meta, eval_path)
        # sanitize
        name = re.sub(r"[^\w.\-가-힣]+", "_", name)
        if not name.endswith(".xlsx"):
            name += ".xlsx"
        out_dir = _artifacts_dir(args.user) / "reports"
        out_path = out_dir / name

    requested_path = out_path
    out_path = _unique_path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(data)
    wb.save(out_path)

    counts = _count_verdicts(data["results"])
    result = {
        "ok": True,
        "path": str(out_path),
        "filename": out_path.name,
        "requested_name": requested_path.name,
        "renamed": out_path.name != requested_path.name,
        "bytes": out_path.stat().st_size,
        "sheets": wb.sheetnames,
        "verdicts": dict(counts),
        "result_rows": len(data["results"]),
        "action_rows": len(data.get("actions") or []),
        "coverage_rows": len(data.get("coverage") or []),
        "fill_check": {
            "meta_target_doc": bool(data["meta"].get("target_doc")),
            "meta_scope": bool(data["meta"].get("scope")),
            "summary_strengths": len(data["summary"].get("strengths") or []),
            "summary_gaps": len(data["summary"].get("gaps") or []),
            "summary_needs": len(data["summary"].get("needs_verification") or []),
            "results_missing_evidence": sum(
                1
                for r in data["results"]
                if not str(r.get("판정근거") or "").strip()
            ),
            "actions_missing_fields": sum(
                1
                for a in (data.get("actions") or [])
                if not (
                    str(a.get("요구사항") or "").strip()
                    and str(a.get("현재상태") or "").strip()
                    and str(a.get("권고조치") or "").strip()
                )
            ),
            "coverage_missing_fields": sum(
                1
                for c in (data.get("coverage") or [])
                if not (
                    str(c.get("섹션") or "").strip()
                    and str(c.get("관련조항") or "").strip()
                )
            ),
        },
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
